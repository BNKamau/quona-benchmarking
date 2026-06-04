"""
Label-based Excel parsers for quarterly/monthly KPI uploads.

Each parser returns a list of dicts keyed to kpi_snapshots columns.
All monetary values are already converted to USD.
"""

import calendar
import io
import re
from datetime import date, datetime

import openpyxl

FX_ZAR: float = 16.5
FX_NGN: float = 1600.0   # NGN/USD — update periodically

SUPPORTED_COMPANIES: set[str] = {"Yoco", "Lulalend", "Verto", "VertoFX", "MaxSoko", "Cowrywise", "Twinco", "TWINCO", "Khazna", "Enza", "SAVA"}


# ── Shared helpers ─────────────────────────────────────────────────────────────

def safe_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", ""))
    except (ValueError, TypeError):
        return None


def to_month_end(year: int, month: int) -> str:
    last = calendar.monthrange(year, month)[1]
    return date(year, month, last).isoformat()


def _normalize_month_str(raw: str) -> str | None:
    """'January 2024', 'Jan 2024', 'December  2025' → YYYY-MM-DD (month end)."""
    raw = re.sub(r"\s+", " ", raw.strip())
    for fmt in ("%B %Y", "%b %Y"):
        try:
            d = datetime.strptime(raw, fmt)
            return to_month_end(d.year, d.month)
        except ValueError:
            pass
    return None


def _parse_pl_header(hdr: str) -> str | None:
    """'Dec-25', 'Jan-26', 'Sept-25' → YYYY-MM-DD (month end)."""
    m = re.fullmatch(r"([A-Za-z]+)-(\d{2,4})", hdr.strip())
    if not m:
        return None
    try:
        month_num = datetime.strptime(m.group(1)[:3], "%b").month
    except ValueError:
        return None
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return to_month_end(year, month_num)


def find_row(
    ws,
    label: str,
    label_col: int = 1,
    max_rows: int = 300,
    exact: bool = False,
) -> int | None:
    """Return first 1-indexed row where cell(row, label_col) matches label."""
    needle = label.lower().strip()
    limit  = min(max_rows + 1, ws.max_row + 1)
    for r in range(1, limit):
        val = ws.cell(r, label_col).value
        if val is None:
            continue
        hay = str(val).strip().lower()
        if exact:
            if hay == needle:
                return r
        else:
            if needle in hay:
                return r
    return None


# ── Yoco ──────────────────────────────────────────────────────────────────────

def _first_row(*labels, ws, label_col: int = 1, exact: bool = False):
    """Return the first matching row for any of the given labels, or None."""
    for lbl in labels:
        r = find_row(ws, lbl, label_col=label_col, exact=exact)
        if r is not None:
            return r
    return None


def parse_yoco(file_bytes: bytes) -> list[dict]:
    """
    Sheet : KPIQuona Export Doc
    Row 1 : metric labels in col A, date strings from col B onward
    Currency: ZAR / 16.5 → USD

    Derived metrics computed per period:
      gross_margin_pct   = gross_profit / revenue          (or revenue - COGS if GP row absent)
      ebitda_margin_pct  = ebitda / revenue
      net_margin_pct     = net_income / revenue
      revenue_growth_pct = (rev - prior_rev) / prior_rev  (None for first period in batch)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "KPIQuona Export Doc" not in wb.sheetnames:
        raise ValueError(
            f"Sheet 'KPIQuona Export Doc' not found. Sheets: {wb.sheetnames}"
        )
    ws = wb["KPIQuona Export Doc"]

    # ── Raw metric row discovery ──────────────────────────────────────────────
    row_rev = find_row(ws, "Transaction Revenue",      label_col=1)
    row_gp  = find_row(ws, "Transaction Gross Margin", label_col=1)
    row_gmv = find_row(ws, "Transaction Volume",       label_col=1)
    row_eop = find_row(ws, "End of Period Base",       label_col=1)
    row_mam = find_row(ws, "Monthly Active Merchants", label_col=1)

    # EBITDA: match "EBITDA - Actuals" (actual label in file) first, then generic variants
    row_ebitda = _first_row(
        "EBITDA - Actuals", "EBITDA - Actual",
        ws=ws, label_col=1, exact=True,
    ) or _first_row(
        "Adjusted EBITDA", "Total EBITDA", "Group EBITDA", "EBITDA",
        ws=ws, label_col=1, exact=True,
    )

    # Net income: match "Net Profit - Actuals" (actual label in file) first, then generic variants
    row_net = _first_row(
        "Net Profit - Actuals", "Net Profit - Actual", "Net Income - Actuals",
        ws=ws, label_col=1, exact=True,
    ) or _first_row(
        "Net Income", "Net Profit", "PAT", "Profit After Tax",
        "Net Profit/(Loss)", "Net Loss", "Profit / (Loss) After Tax",
        ws=ws, label_col=1, exact=True,
    )

    # COGS fallback — only used when Transaction Gross Margin row is absent
    row_cogs = None
    if row_gp is None:
        row_cogs = _first_row(
            "Cost of Sales - Actuals", "Cost of Sales - Actual",
            "Transaction Costs", "Cost of Revenue", "Cost of Goods Sold", "COGS",
            ws=ws, label_col=1, exact=True,
        )

    if row_rev is None:
        raise ValueError(
            "Cannot find 'Transaction Revenue' row in KPIQuona Export Doc sheet"
        )

    # ── Date column discovery (row 1, col B onward) ───────────────────────────
    date_cols: dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        raw = ws.cell(1, c).value
        if isinstance(raw, str) and re.search(r"20\d{2}", raw):
            d = _normalize_month_str(raw)
            if d and d >= "2023-01-01":
                date_cols[c] = d

    # ── Per-period extraction ─────────────────────────────────────────────────
    results: list[dict] = []
    for col, period in sorted(date_cols.items(), key=lambda x: x[1]):
        rev_zar = safe_float(ws.cell(row_rev, col).value)
        if not rev_zar:
            continue

        gp_zar     = safe_float(ws.cell(row_gp,    col).value) if row_gp    else None
        gmv_zar    = safe_float(ws.cell(row_gmv,   col).value) if row_gmv   else None
        eop        = safe_float(ws.cell(row_eop,   col).value) if row_eop   else None
        mam        = safe_float(ws.cell(row_mam,   col).value) if row_mam   else None
        ebitda_zar = safe_float(ws.cell(row_ebitda,col).value) if row_ebitda else None
        net_zar    = safe_float(ws.cell(row_net,   col).value) if row_net   else None

        # Gross profit fallback: Revenue - COGS if GP row not found
        if gp_zar is None and row_cogs is not None:
            cogs_zar = safe_float(ws.cell(row_cogs, col).value)
            if cogs_zar is not None:
                gp_zar = rev_zar - cogs_zar

        # Derived margins (null if inputs unavailable)
        rev_usd      = round(rev_zar / FX_ZAR, 2)
        gp_usd       = round(gp_zar    / FX_ZAR, 2) if gp_zar    is not None else None
        gm_pct       = round(gp_zar    / rev_zar * 100, 4) if gp_zar    is not None else None
        ebitda_usd   = round(ebitda_zar / FX_ZAR, 2) if ebitda_zar is not None else None
        ebitda_m_pct = round(ebitda_zar / rev_zar * 100, 4) if ebitda_zar is not None else None
        net_usd      = round(net_zar    / FX_ZAR, 2) if net_zar    is not None else None
        net_m_pct    = round(net_zar    / rev_zar * 100, 4) if net_zar    is not None else None

        # Revenue growth vs immediately prior period in this batch
        if results and results[-1]["revenue_usd"]:
            prior = results[-1]["revenue_usd"]
            rev_growth = round((rev_usd - prior) / prior * 100, 4) if prior > 0 else None
        else:
            rev_growth = None  # first period in batch; backfilled by _recompute_growth

        results.append({
            "period_end_date":      period,
            "reporting_currency":   "ZAR",
            "fx_rate_to_usd":       FX_ZAR,
            "revenue_usd":          rev_usd,
            "gross_profit_usd":     gp_usd,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           ebitda_usd,
            "ebitda_margin_pct":    ebitda_m_pct,
            "net_income_usd":       net_usd,
            "net_margin_pct":       net_m_pct,
            "revenue_growth_pct":   rev_growth,
            "gmv_usd":              round(gmv_zar / FX_ZAR, 2) if gmv_zar is not None else None,
            "customer_count":       int(eop) if eop else None,
            "active_clients_count": int(mam) if mam else None,
        })

    return results


# ── Verto ─────────────────────────────────────────────────────────────────────

def parse_verto(file_bytes: bytes) -> list[dict]:
    """
    Sheet : KPIs
    Row 2 : datetime date headers from col 3 onward
    Col 1 : metric labels
    Currency: USD (no conversion)
    LTM   : rolling 12 months (all months in file, caller computes LTM)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "KPIs" not in wb.sheetnames:
        raise ValueError(f"Sheet 'KPIs' not found. Sheets: {wb.sheetnames}")
    ws = wb["KPIs"]

    row_rev = find_row(ws, "Total Revenue",                    label_col=1)
    row_gp  = find_row(ws, "Gross Profit",                     label_col=1)
    row_gm  = find_row(ws, "Gross Margin",                     label_col=1, exact=True)
    row_ebt = find_row(ws, "EBITDA",                           label_col=1, exact=True)
    row_tpv = find_row(ws, "Total Processed Volume - Overall", label_col=1)
    row_mac = find_row(ws, "1-month active clients",           label_col=1)

    if row_rev is None:
        raise ValueError("Cannot find 'Total Revenue' row in Verto KPIs sheet")

    # Date columns from row 2 (datetime objects)
    date_cols: dict[int, str] = {}
    for c in range(3, ws.max_column + 1):
        raw = ws.cell(2, c).value
        if hasattr(raw, "year"):
            date_cols[c] = to_month_end(raw.year, raw.month)

    results: list[dict] = []
    for col, period in sorted(date_cols.items(), key=lambda x: x[1]):
        rev = safe_float(ws.cell(row_rev, col).value)
        if not rev:
            continue

        gp  = safe_float(ws.cell(row_gp,  col).value) if row_gp  else None
        gm  = safe_float(ws.cell(row_gm,  col).value) if row_gm  else None
        ebt = safe_float(ws.cell(row_ebt, col).value) if row_ebt else None
        tpv = safe_float(ws.cell(row_tpv, col).value) if row_tpv else None
        mac = safe_float(ws.cell(row_mac, col).value) if row_mac else None

        gm_pct = round(gm  * 100,     4) if gm  is not None else None
        em_pct = round(ebt / rev * 100, 4) if ebt is not None else None

        results.append({
            "period_end_date":      period,
            "reporting_currency":   "USD",
            "fx_rate_to_usd":       1.0,
            "revenue_usd":          round(rev, 2),
            "gross_profit_usd":     round(gp,  2) if gp  is not None else None,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           round(ebt, 2) if ebt is not None else None,
            "ebitda_margin_pct":    em_pct,
            "tpv_usd":              round(tpv, 2) if tpv is not None else None,
            "active_clients_count": int(mac)       if mac is not None else None,
        })

    return results


# ── Lulalend ──────────────────────────────────────────────────────────────────

def _quarter_gp_zar(quarter_end: str, monthly_gp: dict[str, float]) -> float | None:
    """Sum gross profit (ZAR) for the 3 calendar months ending at quarter_end."""
    d = datetime.strptime(quarter_end, "%Y-%m-%d")
    months = []
    for offset in (2, 1, 0):
        m = d.month - offset
        y = d.year
        while m <= 0:
            m += 12
            y -= 1
        months.append(to_month_end(y, m))
    if not all(k in monthly_gp for k in months):
        return None
    return sum(monthly_gp[k] for k in months)


def parse_lulalend(file_bytes: bytes) -> list[dict]:
    import calendar
    from datetime import date

    MONTH_MAP = {
        "january": 1, "february": 2, "march": 3, "april": 4,
        "may": 5, "june": 6, "july": 7, "august": 8,
        "september": 9, "october": 10, "november": 11, "december": 12,
    }

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["KPIs"] if "KPIs" in wb.sheetnames else wb["KPI's"]
    rows = list(ws.iter_rows(max_row=30, values_only=True))

    # Row 5 (index 4): period labels — "March 2026", "December 2025" etc
    # Data columns start at index 3 (column D)
    period_row = rows[4]

    # Build label -> row index map from col C (index 2)
    label_map = {}
    for i, row in enumerate(rows):
        label = row[2]
        if isinstance(label, str):
            label_map[label.strip()] = i

    def _get(label, col_idx):
        i = label_map.get(label)
        if i is None:
            return None
        val = rows[i][col_idx]
        return float(val) if isinstance(val, (int, float)) else None

    ZAR_TO_USD = 18.5

    results = []
    for col_idx in range(3, len(period_row)):
        period_label = period_row[col_idx]
        if not isinstance(period_label, str) or not period_label.strip():
            continue

        parts = period_label.strip().split()
        if len(parts) != 2:
            continue
        month_name, year_str = parts[0].lower().strip(), parts[1].strip()
        month_num = MONTH_MAP.get(month_name)
        if not month_num:
            continue
        try:
            year = int(year_str)
        except ValueError:
            continue
        last_day = calendar.monthrange(year, month_num)[1]
        period_end = date(year, month_num, last_day).strftime("%Y-%m-%d")

        revenue_zar    = _get("Credit Revenue", col_idx)
        ebitda_zar     = _get("EBITDA", col_idx)
        loan_book_zar  = _get("Net Loan Portfolio", col_idx)
        net_yield_raw  = _get("Average Annualized Interest Rate", col_idx)
        par_30_raw     = _get("Par 30 + Restructured loans", col_idx)
        par_90_raw     = _get("Par 90", col_idx)
        active_clients = _get("Total active clients", col_idx)
        unique_smes    = _get("Number of SMEs - Unique to date", col_idx)

        if revenue_zar is None or revenue_zar == 0:
            continue

        revenue_usd   = revenue_zar / ZAR_TO_USD
        ebitda_usd    = ebitda_zar / ZAR_TO_USD if ebitda_zar is not None else None
        loan_book_usd = loan_book_zar / ZAR_TO_USD if loan_book_zar is not None else None
        ebitda_margin = round(ebitda_usd / revenue_usd * 100, 4) if ebitda_usd is not None else None

        net_yield_pct = round(net_yield_raw * 100, 4) if net_yield_raw is not None else None
        par_30_pct    = round(par_30_raw * 100, 4) if par_30_raw is not None else None
        par_90_pct    = round(par_90_raw * 100, 4) if par_90_raw is not None else None

        results.append({
            "period_end_date":        period_end,
            "reporting_currency":     "ZAR",
            "fx_rate_to_usd":         ZAR_TO_USD,
            "revenue_usd":            round(revenue_usd, 2),
            "ebitda_usd":             round(ebitda_usd, 2) if ebitda_usd is not None else None,
            "ebitda_margin_pct":      ebitda_margin,
            "gross_margin_pct":       None,
            "loan_book_gross_usd":    round(loan_book_usd, 2) if loan_book_usd is not None else None,
            "net_yield_pct":          net_yield_pct,
            "par_30_pct":             par_30_pct,
            "par_90_pct":             par_90_pct,
            "active_clients_count":   int(active_clients) if active_clients is not None else None,
            "unique_borrowers_count": int(unique_smes) if unique_smes is not None else None,
            "revenue_growth_pct":     None,
        })

    return results


# ── MaxSoko ───────────────────────────────────────────────────────────────────

def parse_maxsoko(file_bytes: bytes) -> list[dict]:
    """
    Sheet : 'Consolidated View ' (found by prefix, trailing space tolerated)
    Row 4 : datetime date headers; first occurrence of each (year, month) used
    Col 3 : metric labels; data in the same columns as row-4 dates
    Currency: USD (no conversion)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    sheet_name = next(
        (n for n in wb.sheetnames if "consolidated view" in n.lower()), None
    )
    if sheet_name is None:
        raise ValueError(
            f"Cannot find 'Consolidated View' sheet. Sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    # Labels in col C (col 3)
    row_rev = find_row(ws, "Total Revenues", label_col=3)
    row_gp  = find_row(ws, "Gross Profit",   label_col=3, exact=True)
    row_gm  = find_row(ws, "Gross Profit Margin (%)", label_col=3)
    row_ebt = find_row(ws, "EBITDA",         label_col=3, exact=True)
    row_net = (
        _first_row(
            "net income", "net profit", "profit after tax", "pat",
            "net profit/(loss)", "net income/(loss)", "net loss",
            ws=ws, label_col=3, exact=True,
        ) or _first_row(
            "net income", "net profit", "profit after tax",
            ws=ws, label_col=3,
        )
    )
    row_gmv = _first_row(
        "gmv", "gross merchandise value", "total volume", "order volume",
        "total order value", "orders volume", "total orders value",
        ws=ws, label_col=3,
    )
    row_ac = _first_row(
        "active clients", "active buyers", "active users", "active retailers",
        "number of clients", "total clients", "buyers", "retailers",
        ws=ws, label_col=3,
    )

    if row_rev is None:
        raise ValueError(
            "Cannot find 'Total Revenues' row (col C) in MaxSoko Consolidated View"
        )

    # Date columns from row 4; skip duplicate year-month (quarterly summaries)
    seen_ym: set = set()
    date_cols: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(4, c).value
        if not isinstance(raw, datetime):
            continue
        ym = (raw.year, raw.month)
        if ym in seen_ym:
            continue
        seen_ym.add(ym)
        date_cols[c] = to_month_end(raw.year, raw.month)

    results: list[dict] = []
    for col, period in sorted(date_cols.items(), key=lambda x: x[1]):
        rev = safe_float(ws.cell(row_rev, col).value)
        if not rev:
            continue

        gp  = safe_float(ws.cell(row_gp,  col).value) if row_gp  else None
        gm  = safe_float(ws.cell(row_gm,  col).value) if row_gm  else None
        ebt = safe_float(ws.cell(row_ebt, col).value) if row_ebt else None
        net = safe_float(ws.cell(row_net, col).value) if row_net else None
        gmv = safe_float(ws.cell(row_gmv, col).value) if row_gmv else None
        ac  = safe_float(ws.cell(row_ac,  col).value) if row_ac  else None

        # Prefer explicit GM% column; fall back to GP/Rev
        gm_pct = (
            round(gm * 100, 4)          if gm  is not None else
            round(gp / rev * 100, 4)    if gp  is not None else None
        )
        em_pct  = round(ebt / rev * 100, 4) if ebt is not None else None
        net_pct = round(net / rev * 100, 4) if net is not None else None

        if results and results[-1].get("revenue_usd"):
            prior = results[-1]["revenue_usd"]
            rev_growth = round((rev - prior) / prior * 100, 4) if prior > 0 else None
        else:
            rev_growth = None

        results.append({
            "period_end_date":      period,
            "reporting_currency":   "USD",
            "fx_rate_to_usd":       1.0,
            "revenue_usd":          round(rev, 2),
            "gross_profit_usd":     round(gp,  2) if gp  is not None else None,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           round(ebt, 2) if ebt is not None else None,
            "ebitda_margin_pct":    em_pct,
            "net_income_usd":       round(net, 2) if net is not None else None,
            "net_margin_pct":       net_pct,
            "gmv_usd":              round(gmv, 2) if gmv is not None else None,
            "active_clients_count": int(ac) if ac is not None else None,
            "revenue_growth_pct":   rev_growth,
        })

    return results


# ── Cowrywise ─────────────────────────────────────────────────────────────────

def _cwry_best_sheet(wb):
    """Return the worksheet most likely to contain KPI / P&L data."""
    keywords = ("kpi", "financial", "p&l", "income", "dashboard", "summary", "monthly", "data")
    for kw in keywords:
        for name in wb.sheetnames:
            if kw in name.lower():
                return wb[name]
    return wb.active


def _cwry_label_col(ws) -> int:
    """Return the column (1-indexed) that contains the most metric-label text."""
    best_col, best_n = 1, 0
    for col in (1, 2, 3):
        n = 0
        for r in range(1, 60):
            v = ws.cell(r, col).value
            if isinstance(v, str) and len(v.strip()) > 3 and not re.match(r"^[\d.,\s%\-]+$", v.strip()):
                n += 1
        if n > best_n:
            best_n, best_col = n, col
    return best_col


def _cwry_currency(ws, label_col: int) -> tuple[float, str]:
    """Detect reporting currency; defaults to NGN for Cowrywise."""
    for r in range(1, 20):
        v = str(ws.cell(r, label_col).value or "").lower()
        if "usd" in v or "dollar" in v:
            return 1.0, "USD"
    return FX_NGN, "NGN"


def _cwry_date_cols(ws, label_col: int) -> dict[int, str]:
    """
    Scan rows 1–6 for date-like column headers; return {col: 'YYYY-MM-DD'}.
    Handles datetime objects, 'Jan 2024' / 'January 2024', and 'Dec-24' strings.
    Uses the first header row that yields at least 2 date columns.
    """
    for hrow in range(1, 7):
        found: dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            if c == label_col:
                continue
            raw = ws.cell(hrow, c).value
            if raw is None:
                continue
            if hasattr(raw, "year") and hasattr(raw, "month"):
                found[c] = to_month_end(raw.year, raw.month)
            elif isinstance(raw, str):
                d = _normalize_month_str(raw.strip()) or _parse_pl_header(raw.strip())
                if d:
                    found[c] = d
        if len(found) >= 2:
            return found
    return {}


def parse_cowrywise(file_bytes: bytes) -> list[dict]:
    """
    Flexible label-based parser for Cowrywise Excel exports.

    Searches any sheet for common label variants of each KPI rather than
    relying on a fixed layout, since no canonical file format exists yet.

    Currency detection: searches sheet headers for NGN/USD indicator;
    defaults to NGN (FX_NGN / USD) when ambiguous.

    Metrics extracted:
      revenue_usd, gross_margin_pct, ebitda_usd, ebitda_margin_pct,
      net_income_usd, net_margin_pct, aum_usd, active_clients_count,
      revenue_growth_pct
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _cwry_best_sheet(wb)
    lc = _cwry_label_col(ws)
    fx, currency = _cwry_currency(ws, lc)

    def _find(*labels, exact=False):
        return _first_row(*labels, ws=ws, label_col=lc, exact=exact)

    # Revenue — many possible label variants
    row_rev = _find(
        "total revenue", "revenue", "net revenue", "total net revenue",
        "revenues", "total revenues", "income", "total income",
    )

    # Gross profit (used as fallback when no explicit GM% row exists)
    row_gp = _find("gross profit", "gross profit amount", "gross income")

    # Gross margin percentage (preferred over GP/Rev computation)
    row_gm = _find(
        "gross margin %", "gross margin pct", "gross margin percentage",
        "gross profit margin", "gross profit margin %", "gross profit %",
        "gross margin",
    )

    # EBITDA — exact match first to avoid matching "EBITDA Margin"
    row_ebt = (
        _find("ebitda", exact=True)
        or _find("adjusted ebitda", "total ebitda", "ebitda (loss)", "ebitda profit")
        or _find("ebitda")
    )

    # Net income — exact match first to avoid matching "Net Income Margin"
    row_net = (
        _find(
            "net income", "net profit", "profit after tax", "pat",
            "net profit/(loss)", "net income/(loss)", "net loss",
            "profit/(loss) after tax",
            exact=True,
        )
        or _find("net income", "net profit", "profit after tax")
    )

    # AUM
    row_aum = _find(
        "aum", "total aum", "assets under management",
        "total assets under management", "aum (ngn)", "aum (usd)",
        "funds under management", "fum",
    )

    # Active clients / users
    row_ac = _find(
        "active users", "active clients", "monthly active users", "mau",
        "active investors", "total active users", "registered users",
        "customers", "total customers", "number of users", "total users",
    )

    if row_rev is None:
        raise ValueError(
            f"Cannot find a revenue row in sheet '{ws.title}' (label col {lc}). "
            "Expected labels such as 'Revenue', 'Total Revenue', or 'Net Revenue'."
        )

    date_cols = _cwry_date_cols(ws, lc)
    if not date_cols:
        raise ValueError(
            f"No date columns found in sheet '{ws.title}'. "
            "Expected month/year headers in rows 1–6."
        )

    results: list[dict] = []
    for col, period in sorted(date_cols.items(), key=lambda x: x[1]):

        def _g(row, _col=col):
            return safe_float(ws.cell(row, _col).value) if row else None

        rev_raw = _g(row_rev)
        if not rev_raw:
            continue

        gp_raw  = _g(row_gp)
        gm_raw  = _g(row_gm)
        ebt_raw = _g(row_ebt)
        net_raw = _g(row_net)
        aum_raw = _g(row_aum)
        ac_raw  = _g(row_ac)

        rev_usd = round(rev_raw / fx, 2)

        # Gross margin: prefer explicit % row; accept decimal (0.xx) or percentage (xx)
        if gm_raw is not None:
            gm_pct = round(gm_raw * 100, 4) if abs(gm_raw) <= 1.5 else round(float(gm_raw), 4)
        elif gp_raw is not None:
            gm_pct = round(gp_raw / rev_raw * 100, 4)
        else:
            gm_pct = None

        ebt_usd = round(ebt_raw / fx, 2) if ebt_raw is not None else None
        em_pct  = round(ebt_raw / rev_raw * 100, 4) if ebt_raw is not None else None

        net_usd = round(net_raw / fx, 2) if net_raw is not None else None
        nm_pct  = round(net_raw / rev_raw * 100, 4) if net_raw is not None else None

        aum_usd = round(aum_raw / fx, 2) if aum_raw is not None else None

        if results and results[-1].get("revenue_usd"):
            prior = results[-1]["revenue_usd"]
            rev_growth = round((rev_usd - prior) / prior * 100, 4) if prior > 0 else None
        else:
            rev_growth = None

        results.append({
            "period_end_date":      period,
            "reporting_currency":   currency,
            "fx_rate_to_usd":       fx,
            "revenue_usd":          rev_usd,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           ebt_usd,
            "ebitda_margin_pct":    em_pct,
            "net_income_usd":       net_usd,
            "net_margin_pct":       nm_pct,
            "aum_usd":              aum_usd,
            "active_clients_count": int(ac_raw) if ac_raw is not None else None,
            "revenue_growth_pct":   rev_growth,
        })

    return results


# ── Twinco ────────────────────────────────────────────────────────────────────

def parse_twinco(file_bytes: bytes) -> list[dict]:
    """
    Flexible label-based parser for Twinco Excel exports.

    Auto-detects sheet, label column, currency, and date columns (same heuristics
    as parse_cowrywise) since no canonical file format exists yet.

    Metrics extracted:
      revenue_usd, gross_margin_pct, ebitda_usd, ebitda_margin_pct,
      net_income_usd, net_margin_pct, gmv_usd (volume financed),
      active_clients_count, revenue_growth_pct
    Currency defaults to USD; detects EUR/GBP from header cells.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    ws = _cwry_best_sheet(wb)
    lc = _cwry_label_col(ws)

    # Currency detection — Twinco is typically USD-denominated
    fx, currency = 1.0, "USD"
    for r in range(1, 20):
        v = str(ws.cell(r, lc).value or "").lower()
        if "eur" in v or "euro" in v:
            currency = "EUR"
            break
        if "gbp" in v:
            currency = "GBP"
            break

    def _find(*labels, exact=False):
        return _first_row(*labels, ws=ws, label_col=lc, exact=exact)

    row_rev = _find(
        "total revenue", "revenue", "net revenue", "total net revenue",
        "revenues", "total revenues", "financing income", "total financing income",
        "interest income", "income",
    )
    row_gp = _find("gross profit", "gross income", "net interest income")
    row_gm = _find(
        "gross margin %", "gross margin pct", "gross margin percentage",
        "gross profit margin", "gross profit margin %", "gross profit %",
        "gross margin",
    )
    row_ebt = (
        _find("ebitda", exact=True)
        or _find("adjusted ebitda", "total ebitda", "ebitda (loss)", "ebitda profit")
        or _find("ebitda")
    )
    row_net = (
        _find(
            "net income", "net profit", "profit after tax", "pat",
            "net profit/(loss)", "net income/(loss)", "net loss",
            "profit/(loss) after tax",
            exact=True,
        )
        or _find("net income", "net profit", "profit after tax")
    )
    # Volume financed — PO / SCF / transaction volume → stored in gmv_usd
    row_vol = _find(
        "volume financed", "total volume financed", "po volume", "po financed",
        "purchase order volume", "purchase orders financed", "total po",
        "volume of purchase orders", "financed volume", "total financed volume",
        "transaction volume", "total transaction volume", "gmv",
    )
    row_ac = _find(
        "active clients", "active buyers", "active suppliers", "active counterparties",
        "number of clients", "total clients", "buyers", "suppliers",
        "clients", "counterparties", "active companies",
    )

    if row_rev is None:
        raise ValueError(
            f"Cannot find a revenue row in sheet '{ws.title}' (label col {lc}). "
            "Expected labels such as 'Revenue', 'Total Revenue', or 'Financing Income'."
        )

    date_cols = _cwry_date_cols(ws, lc)
    if not date_cols:
        raise ValueError(
            f"No date columns found in sheet '{ws.title}'. "
            "Expected month/year headers in rows 1–6."
        )

    results: list[dict] = []
    for col, period in sorted(date_cols.items(), key=lambda x: x[1]):

        def _g(row, _col=col):
            return safe_float(ws.cell(row, _col).value) if row else None

        rev_raw = _g(row_rev)
        if not rev_raw:
            continue

        gp_raw  = _g(row_gp)
        gm_raw  = _g(row_gm)
        ebt_raw = _g(row_ebt)
        net_raw = _g(row_net)
        vol_raw = _g(row_vol)
        ac_raw  = _g(row_ac)

        rev_usd = round(rev_raw / fx, 2)

        if gm_raw is not None:
            gm_pct = round(gm_raw * 100, 4) if abs(gm_raw) <= 1.5 else round(float(gm_raw), 4)
        elif gp_raw is not None:
            gm_pct = round(gp_raw / rev_raw * 100, 4)
        else:
            gm_pct = None

        ebt_usd = round(ebt_raw / fx, 2) if ebt_raw is not None else None
        em_pct  = round(ebt_raw / rev_raw * 100, 4) if ebt_raw is not None else None

        net_usd = round(net_raw / fx, 2) if net_raw is not None else None
        nm_pct  = round(net_raw / rev_raw * 100, 4) if net_raw is not None else None

        vol_usd = round(vol_raw / fx, 2) if vol_raw is not None else None

        if results and results[-1].get("revenue_usd"):
            prior = results[-1]["revenue_usd"]
            rev_growth = round((rev_usd - prior) / prior * 100, 4) if prior > 0 else None
        else:
            rev_growth = None

        results.append({
            "period_end_date":      period,
            "reporting_currency":   currency,
            "fx_rate_to_usd":       fx,
            "revenue_usd":          rev_usd,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           ebt_usd,
            "ebitda_margin_pct":    em_pct,
            "net_income_usd":       net_usd,
            "net_margin_pct":       nm_pct,
            "gmv_usd":              vol_usd,
            "active_clients_count": int(ac_raw) if ac_raw is not None else None,
            "revenue_growth_pct":   rev_growth,
        })

    return results


# ── Khazna ────────────────────────────────────────────────────────────────────

def parse_khazna(file_bytes: bytes) -> list[dict]:
    """
    Flexible label-based parser for Khazna Excel exports.

    Auto-detects sheet, label column, currency, and date columns.
    Currency defaults to USD (Quona typically receives USD-denominated reports).

    Metrics extracted:
      revenue_usd, gross_margin_pct, ebitda_usd, ebitda_margin_pct,
      net_income_usd, net_margin_pct, arr_usd, loan_book_gross_usd,
      par_90_pct, active_clients_count, revenue_growth_pct
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    ws = _cwry_best_sheet(wb)
    lc = _cwry_label_col(ws)

    # Currency detection — default USD
    fx, currency = 1.0, "USD"
    for r in range(1, 20):
        v = str(ws.cell(r, lc).value or "").lower()
        if "egp" in v or "egyptian pound" in v:
            fx, currency = 30.9, "EGP"
            break
        if "sar" in v or "riyal" in v:
            fx, currency = 3.75, "SAR"
            break

    def _find(*labels, exact=False):
        return _first_row(*labels, ws=ws, label_col=lc, exact=exact)

    row_rev = _find(
        "total revenue", "revenue", "net revenue", "total net revenue",
        "revenues", "interest income", "financing income", "total income",
    )
    row_gp = _find("gross profit", "gross income")
    row_gm = _find(
        "gross margin %", "gross margin pct", "gross margin percentage",
        "gross profit margin", "gross profit %", "gross margin",
    )
    row_ebt = (
        _find("ebitda", exact=True)
        or _find("adjusted ebitda", "total ebitda", "ebitda (loss)")
        or _find("ebitda")
    )
    row_net = (
        _find(
            "net income", "net profit", "profit after tax", "pat",
            "net profit/(loss)", "net income/(loss)", "net loss",
            exact=True,
        )
        or _find("net income", "net profit", "profit after tax")
    )
    # ARR — annualised recurring revenue
    row_arr = _find(
        "arr", "annualised recurring revenue", "annual recurring revenue",
        "annualised revenue", "annualized revenue",
    )
    # Loan book
    row_lb = _find(
        "loan book", "net loan book", "gross loan book", "loan portfolio",
        "net loan portfolio", "gross loan portfolio", "outstanding loans",
        "total loans", "loan receivables",
    )
    # PAR 90
    row_par90 = _find(
        "par 90", "par90", "par-90", "portfolio at risk 90",
        "npl", "npl rate", "non-performing loans",
        exact=True,
    ) or _find("par 90", "par90", "npl")
    # Active clients / workers
    row_ac = _find(
        "active clients", "active workers", "active employees", "active users",
        "number of active clients", "total active clients", "registered workers",
        "active borrowers", "enrolled workers",
    )

    if row_rev is None:
        raise ValueError(
            f"Cannot find a revenue row in sheet '{ws.title}' (label col {lc}). "
            "Expected labels such as 'Revenue', 'Total Revenue', or 'Interest Income'."
        )

    date_cols = _cwry_date_cols(ws, lc)
    if not date_cols:
        raise ValueError(
            f"No date columns found in sheet '{ws.title}'. "
            "Expected month/year headers in rows 1–6."
        )

    results: list[dict] = []
    for col, period in sorted(date_cols.items(), key=lambda x: x[1]):

        def _g(row, _col=col):
            return safe_float(ws.cell(row, _col).value) if row else None

        rev_raw = _g(row_rev)
        if not rev_raw:
            continue

        gp_raw  = _g(row_gp)
        gm_raw  = _g(row_gm)
        ebt_raw = _g(row_ebt)
        net_raw = _g(row_net)
        arr_raw = _g(row_arr)
        lb_raw  = _g(row_lb)
        p90_raw = _g(row_par90)
        ac_raw  = _g(row_ac)

        rev_usd = round(rev_raw / fx, 2)

        if gm_raw is not None:
            gm_pct = round(gm_raw * 100, 4) if abs(gm_raw) <= 1.5 else round(float(gm_raw), 4)
        elif gp_raw is not None:
            gm_pct = round(gp_raw / rev_raw * 100, 4)
        else:
            gm_pct = None

        ebt_usd = round(ebt_raw / fx, 2) if ebt_raw is not None else None
        em_pct  = round(ebt_raw / rev_raw * 100, 4) if ebt_raw is not None else None

        net_usd = round(net_raw / fx, 2) if net_raw is not None else None
        nm_pct  = round(net_raw / rev_raw * 100, 4) if net_raw is not None else None

        arr_usd = round(arr_raw / fx, 2) if arr_raw is not None else None
        lb_usd  = round(lb_raw  / fx, 2) if lb_raw  is not None else None

        # PAR 90: accept decimal (0.004) or percentage (0.4) form
        par90 = None
        if p90_raw is not None:
            par90 = round(p90_raw * 100, 4) if abs(p90_raw) < 1 else round(float(p90_raw), 4)

        if results and results[-1].get("revenue_usd"):
            prior = results[-1]["revenue_usd"]
            rev_growth = round((rev_usd - prior) / prior * 100, 4) if prior > 0 else None
        else:
            rev_growth = None

        results.append({
            "period_end_date":      period,
            "reporting_currency":   currency,
            "fx_rate_to_usd":       fx,
            "revenue_usd":          rev_usd,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           ebt_usd,
            "ebitda_margin_pct":    em_pct,
            "net_income_usd":       net_usd,
            "net_margin_pct":       nm_pct,
            "arr_usd":              arr_usd,
            "loan_book_gross_usd":  lb_usd,
            "par_90_pct":           par90,
            "active_clients_count": int(ac_raw) if ac_raw is not None else None,
            "revenue_growth_pct":   rev_growth,
        })

    return results


def parse_enza(file_bytes: bytes) -> list[dict]:
    import io
    from datetime import datetime
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["KPIs"]

    rows = list(ws.iter_rows(max_row=60, values_only=True))

    date_row = rows[0]

    label_map = {}
    for i, row in enumerate(rows):
        label = row[1]
        if isinstance(label, str):
            label_map[label.strip()] = i

    def _get(label, col_idx):
        i = label_map.get(label)
        if i is None:
            return None
        val = rows[i][col_idx]
        if isinstance(val, str):
            return None
        return val

    results = []
    for col_idx in range(4, len(date_row)):
        date_val = date_row[col_idx]
        if not isinstance(date_val, datetime):
            continue

        period_end = date_val.strftime("%Y-%m-%d")

        revenue = _get("Total Revenues", col_idx)
        if revenue is None or revenue == 0:
            continue

        gp         = _get("Contribution Margin ", col_idx)
        gm_pct_raw = _get("Contibution Margin ", col_idx)  # typo is in the source file
        ebitda     = _get("EBITDA", col_idx)
        clients    = _get("Number of Clients (Banks/Institutions)", col_idx)

        gm_pct = round(float(gm_pct_raw) * 100, 4) if gm_pct_raw is not None else None
        em_pct = round(float(ebitda) / float(revenue) * 100, 4) if (ebitda is not None and revenue) else None

        results.append({
            "period_end_date":      period_end,
            "reporting_currency":   "USD",
            "revenue_usd":          float(revenue),
            "gross_profit_usd":     float(gp) if gp is not None else None,
            "gross_margin_pct":     gm_pct,
            "ebitda_usd":           float(ebitda) if ebitda is not None else None,
            "ebitda_margin_pct":    em_pct,
            "active_clients_count": int(clients) if clients is not None else None,
            "revenue_growth_pct":   None,
        })

    return results


def parse_sava(file_bytes: bytes) -> list[dict]:
    import io
    import calendar
    from datetime import datetime, date
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Revenue To Date"]
    rows = list(ws.iter_rows(max_row=22, values_only=True))

    results = []
    for row in rows:
        date_val = row[0]
        if not isinstance(date_val, datetime):
            continue

        revenue_usd = row[2]
        if not isinstance(revenue_usd, (int, float)) or revenue_usd == 0:
            continue

        revenue_zar = row[1]
        last_day = calendar.monthrange(date_val.year, date_val.month)[1]
        period_end = date(date_val.year, date_val.month, last_day).strftime("%Y-%m-%d")

        fx_rate = None
        if isinstance(revenue_zar, (int, float)) and revenue_zar and revenue_usd:
            fx_rate = round(revenue_zar / revenue_usd, 6)

        results.append({
            "period_end_date":      period_end,
            "reporting_currency":   "ZAR",
            "fx_rate_to_usd":       fx_rate,
            "revenue_usd":          float(revenue_usd),
            "gross_profit_usd":     None,
            "gross_margin_pct":     None,
            "ebitda_usd":           None,
            "ebitda_margin_pct":    None,
            "active_clients_count": None,
            "revenue_growth_pct":   None,
        })

    return results


# ── Registry ──────────────────────────────────────────────────────────────────

PARSERS: dict[str, callable] = {
    "Yoco":      parse_yoco,
    "Lulalend":  parse_lulalend,
    "Verto":     parse_verto,
    "VertoFX":   parse_verto,
    "MaxSoko":   parse_maxsoko,
    "Cowrywise": parse_cowrywise,
    "Twinco":    parse_twinco,
    "TWINCO":    parse_twinco,
    "Khazna":    parse_khazna,
    "Enza":      parse_enza,
    "SAVA":      parse_sava,
}
