"""
Fix missing/incorrect KPI data for Yoco, Lulalend, Khazna, AllLife, OCTA, MaxSoko.
Run from the project root:  python scripts/fix_company_data.py
"""

import calendar
import re
import sqlite3
import warnings
from datetime import date, datetime

import openpyxl

warnings.filterwarnings("ignore")

DB_PATH = "benchmarking.db"
FX_ZAR = 18.5          # ZAR -> USD conversion rate (as instructed)

# ── DB helpers ─────────────────────────────────────────────────────────────────

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row


def to_month_end(val) -> str | None:
    """Normalise any date/datetime/string to the last day of its month."""
    if val is None:
        return None
    if isinstance(val, datetime):
        y, m = val.year, val.month
    elif isinstance(val, date):
        y, m = val.year, val.month
    elif isinstance(val, str):
        val = val.strip()
        # Try common patterns
        for fmt in ("%B %Y", "%b %Y", "%B  %Y"):
            try:
                d = datetime.strptime(val, fmt)
                y, m = d.year, d.month
                break
            except ValueError:
                continue
        else:
            return None
    else:
        return None
    last = calendar.monthrange(y, m)[1]
    return date(y, m, last).isoformat()


def safe_float(v) -> float | None:
    """Parse numbers that may contain spaces or commas (e.g. '50 646 708')."""
    if v is None:
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", ""))
    except (ValueError, TypeError):
        return None


def upsert(data: dict) -> str:
    """
    INSERT or UPDATE a kpi_snapshots row keyed on (company_id, period_end_date).
    For updates only non-None incoming values overwrite existing columns.
    Returns 'inserted' or 'updated'.
    """
    existing = conn.execute(
        "SELECT id FROM kpi_snapshots WHERE company_id=? AND period_end_date=?",
        (data["company_id"], data["period_end_date"]),
    ).fetchone()

    now = datetime.utcnow().isoformat()

    if existing:
        cols = {
            k: v
            for k, v in data.items()
            if k not in ("company_id", "period_end_date") and v is not None
        }
        if cols:
            set_sql = ", ".join(f"{k}=?" for k in cols)
            conn.execute(
                f"UPDATE kpi_snapshots SET {set_sql}, updated_at=?"
                f" WHERE company_id=? AND period_end_date=?",
                [*cols.values(), now, data["company_id"], data["period_end_date"]],
            )
        return "updated"
    else:
        data.setdefault("created_at", now)
        data["updated_at"] = now
        cols = list(data.keys())
        conn.execute(
            f"INSERT INTO kpi_snapshots ({', '.join(cols)})"
            f" VALUES ({', '.join('?' * len(cols))})",
            list(data.values()),
        )
        return "inserted"


# ── Snapshot: row counts & metric coverage per company ─────────────────────────

def snapshot(company_ids: list[int]) -> dict:
    result = {}
    for cid in company_ids:
        row = conn.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(revenue_usd       IS NOT NULL) AS rev,
                SUM(gross_margin_pct  IS NOT NULL) AS gm,
                SUM(ebitda_usd        IS NOT NULL) AS ebitda,
                SUM(loan_book_gross_usd IS NOT NULL) AS loan,
                SUM(par_30_pct        IS NOT NULL) AS par30,
                SUM(net_yield_pct     IS NOT NULL) AS yield_,
                SUM(arr_usd           IS NOT NULL) AS arr,
                SUM(customer_count    IS NOT NULL) AS cust,
                MAX(period_end_date)  AS last_period,
                MAX(CASE WHEN revenue_usd IS NOT NULL
                         THEN revenue_usd END) AS latest_rev
            FROM kpi_snapshots
            WHERE company_id = ?
            """,
            (cid,),
        ).fetchone()
        result[cid] = dict(row)
    return result


COMPANY_IDS = [2, 5, 6, 8, 10, 11]
COMPANY_NAMES = {2: "Yoco", 5: "Lulalend", 6: "Khazna",
                 8: "MaxSoko", 10: "AllLife", 11: "OCTA"}

print("=" * 68)
print("BEFORE")
print("=" * 68)
before = snapshot(COMPANY_IDS)
for cid, s in before.items():
    print(f"  {COMPANY_NAMES[cid]:12s} rows={s['total']:3d}  "
          f"rev={s['rev']}/{s['total']}  gm={s['gm']}/{s['total']}  "
          f"ebitda={s['ebitda']}/{s['total']}  "
          f"loan={s['loan']}/{s['total']}  par30={s['par30']}/{s['total']}  "
          f"yield={s['yield_']}/{s['total']}  arr={s['arr']}/{s['total']}")


# ══════════════════════════════════════════════════════════════════════════════
# 1. YOCO  (company_id = 2)
#    Sheet: KPIQuona Export Doc
#    Row 1: month headers (col A = "Metric Name"; dates start around col 77)
#    Row 12: End of Period Base after churn  -> customer_count
#    Row 13: Monthly Active Merchants        -> active_clients_count
#    Row 16: Transaction Volume              -> gmv_usd   (ZAR -> USD)
#    Row 17: Transaction Revenue excl. VAT  -> revenue_usd (ZAR -> USD)
#    Row 18: Transaction Gross Margin       -> gross_profit_usd + gross_margin_pct
# ══════════════════════════════════════════════════════════════════════════════
print("\n--- Yoco ---")

wb = openpyxl.load_workbook(
    "data/Yoco Quona KPIs_02112026.xlsx", data_only=True
)
ws = wb["KPIQuona Export Doc"]

# Collect date columns that are ≥ 2023-01-01
date_cols: dict[int, str] = {}
for c in range(2, ws.max_column + 1):
    raw = ws.cell(1, c).value
    if isinstance(raw, str) and re.search(r"20[0-9]{2}", raw):
        d = to_month_end(raw)
        if d and d >= "2023-01-01":
            date_cols[c] = d

stats = {"inserted": 0, "updated": 0, "skipped": 0}
for col, period in sorted(date_cols.items(), key=lambda x: x[1]):
    rev_zar = safe_float(ws.cell(17, col).value)
    gp_zar  = safe_float(ws.cell(18, col).value)
    gmv_zar = safe_float(ws.cell(16, col).value)
    eop_base     = safe_float(ws.cell(12, col).value)
    active_merch = safe_float(ws.cell(13, col).value)

    if rev_zar is None or rev_zar == 0:
        stats["skipped"] += 1
        continue

    rev_usd = rev_zar / FX_ZAR
    gm_pct  = (gp_zar / rev_zar * 100) if gp_zar else None

    res = upsert({
        "company_id":         2,
        "period_end_date":    period,
        "reporting_currency": "ZAR",
        "fx_rate_to_usd":     FX_ZAR,
        "revenue_usd":        round(rev_usd, 2),
        "gross_profit_usd":   round(gp_zar / FX_ZAR, 2) if gp_zar else None,
        "gross_margin_pct":   round(gm_pct, 4) if gm_pct else None,
        "gmv_usd":            round(gmv_zar / FX_ZAR, 2) if gmv_zar else None,
        "customer_count":     int(eop_base) if eop_base else None,
        "active_clients_count": int(active_merch) if active_merch else None,
    })
    stats[res] += 1

conn.commit()
print(f"  inserted={stats['inserted']}  updated={stats['updated']}  "
      f"skipped(no rev)={stats['skipped']}  "
      f"date range: {min(date_cols.values())} -> {max(date_cols.values())}")


# ══════════════════════════════════════════════════════════════════════════════
# 2. LULALEND  (company_id = 5)
#    Sheet: KPI's
#    Row 5: quarter end dates (cols 4–16)
#    Row 18: Average Annualized Interest Rate (decimal) -> net_yield_pct
#    All other metrics already seeded correctly; this pass adds the yield.
# ══════════════════════════════════════════════════════════════════════════════
print("\n--- Lulalend ---")

wb = openpyxl.load_workbook(
    "data/Lulalend 12. Investor Man Acc - December 2025 - Quona.xlsx",
    data_only=True,
)
ws = wb["KPI's"]

# Build date -> column map from row 5
lula_date_cols: dict[int, str] = {}
for c in range(4, 20):
    raw = ws.cell(5, c).value
    if isinstance(raw, str) and raw.strip():
        d = to_month_end(raw)
        if d:
            lula_date_cols[c] = d

stats = {"inserted": 0, "updated": 0}
for col, period in lula_date_cols.items():
    yield_raw = ws.cell(18, col).value
    if yield_raw is None:
        continue
    res = upsert({
        "company_id":      5,
        "period_end_date": period,
        "net_yield_pct":   round(float(yield_raw) * 100, 4),
    })
    stats[res] += 1

conn.commit()
print(f"  inserted={stats['inserted']}  updated={stats['updated']}  "
      f"(net_yield_pct added for {sum(stats.values())} quarterly rows)")


# ══════════════════════════════════════════════════════════════════════════════
# 3. KHAZNA  (company_id = 6)
#    All 15 monthly rows already seeded with revenue, ARR, loan book,
#    PAR30, active users, and ebitda_usd.  No data changes required.
#    EBITDA margin now computed in the dashboard from ebitda_usd / revenue_usd.
# ══════════════════════════════════════════════════════════════════════════════
print("\n--- Khazna ---")
n = conn.execute(
    "SELECT COUNT(*) FROM kpi_snapshots WHERE company_id=6"
).fetchone()[0]
print(f"  {n} rows already complete (revenue, ARR, loan book, PAR30, active users, EBITDA). "
      f"No changes needed.")


# ══════════════════════════════════════════════════════════════════════════════
# 4. ALLLIFE  (company_id = 10)
#    8 annual rows (2018–2025) already seeded with revenue_usd, ebitda_usd,
#    and insurance_policies_active.  Gross margin is not in source data.
#    EBITDA margin computed in dashboard from ebitda_usd / revenue_usd.
# ══════════════════════════════════════════════════════════════════════════════
print("\n--- AllLife ---")
n = conn.execute(
    "SELECT COUNT(*) FROM kpi_snapshots WHERE company_id=10"
).fetchone()[0]
ni = conn.execute(
    "SELECT COUNT(*) FROM kpi_snapshots WHERE company_id=10 AND insurance_policies_active IS NOT NULL"
).fetchone()[0]
print(f"  {n} rows, insurance_policies_active populated for {ni} rows. "
      f"Gross margin not available in source. No changes needed.")


# ══════════════════════════════════════════════════════════════════════════════
# 5. OCTA  (company_id = 11)
#    arr_usd already populated for rows Jul 2024 -> Dec 2025.
#    revenue_usd is NULL for all rows -> set revenue_usd = arr_usd
#    Jan 2026: Excel shows '~1.55Mn' (string) -> set arr_usd = 1_550_000 first.
# ══════════════════════════════════════════════════════════════════════════════
print("\n--- OCTA ---")

# Fix Jan 2026 ARR which was a string in the source ('~1.55Mn')
conn.execute(
    """UPDATE kpi_snapshots
       SET arr_usd = 1550000, updated_at = datetime('now')
       WHERE company_id = 11 AND period_end_date = '2026-01-31' AND arr_usd IS NULL"""
)

# Propagate arr_usd -> revenue_usd for all rows that have arr data
conn.execute(
    """UPDATE kpi_snapshots
       SET revenue_usd = arr_usd, updated_at = datetime('now')
       WHERE company_id = 11 AND arr_usd IS NOT NULL AND revenue_usd IS NULL"""
)
conn.commit()

n_rev = conn.execute(
    "SELECT COUNT(*) FROM kpi_snapshots WHERE company_id=11 AND revenue_usd IS NOT NULL"
).fetchone()[0]
n_total = conn.execute(
    "SELECT COUNT(*) FROM kpi_snapshots WHERE company_id=11"
).fetchone()[0]
print(f"  revenue_usd set from arr_usd for {n_rev}/{n_total} rows "
      f"(rows without arr data have no revenue proxy).")


# ══════════════════════════════════════════════════════════════════════════════
# 6. MAXSOKO  (company_id = 8)
#    Source: "Consolidated View " sheet, row 4 = month-end dates (USD),
#    unique monthly columns (one per calendar month, skipping quarterly/annual
#    summary columns that repeat the same year-month).
#      Row 4:  month-end dates (datetime)
#      Row 44: Total Revenues USD          -> revenue_usd
#      Row 52: Gross Profit USD            -> gross_profit_usd
#      Row 53: GP Margin % (decimal)       -> gross_margin_pct
#      Row 10: E-Commerce GMV USD          -> gmv_usd
#      Row 90: EBITDA USD                  -> ebitda_usd
#    Note: the "Redash EG" sheet only contains Egypt-country data; using it
#    previously produced Egypt-only figures (~$145M LTM vs $224.4M consolidated).
# ══════════════════════════════════════════════════════════════════════════════
print("\n--- MaxSoko ---")

import calendar as _cal

wb = openpyxl.load_workbook(
    "data/25 - 12 - MaxSoko Management Accounts.xlsx", data_only=True
)
ws_cv = wb["Consolidated View "]

# Identify unique monthly columns from row 4 (first occurrence of each year-month)
seen_ym: set = set()
cv_date_cols: dict[int, str] = {}
for c in range(1, ws_cv.max_column + 1):
    raw = ws_cv.cell(4, c).value
    if not isinstance(raw, datetime):
        continue
    ym = (raw.year, raw.month)
    if ym in seen_ym:
        continue
    seen_ym.add(ym)
    last_day = _cal.monthrange(raw.year, raw.month)[1]
    cv_date_cols[c] = date(raw.year, raw.month, last_day).isoformat()

stats = {"inserted": 0, "updated": 0, "skipped": 0}
for col, period in sorted(cv_date_cols.items(), key=lambda x: x[1]):
    rev = safe_float(ws_cv.cell(44, col).value)   # Total Revenues USD
    if rev is None or rev == 0:
        stats["skipped"] += 1
        continue
    gp  = safe_float(ws_cv.cell(52, col).value)   # Gross Profit USD
    gm  = safe_float(ws_cv.cell(53, col).value)   # GP Margin % (decimal)
    ebt = safe_float(ws_cv.cell(90, col).value)   # EBITDA USD
    gmv = safe_float(ws_cv.cell(10, col).value)   # E-Commerce GMV USD

    gm_pct = round(gm * 100, 4) if gm is not None else None
    em_pct = round(ebt / rev * 100, 4) if ebt is not None and rev != 0 else None

    res = upsert({
        "company_id":         8,
        "period_end_date":    period,
        "reporting_currency": "USD",
        "fx_rate_to_usd":     1.0,
        "revenue_usd":        round(rev, 2),
        "gross_profit_usd":   round(gp, 2) if gp is not None else None,
        "gross_margin_pct":   gm_pct,
        "gmv_usd":            round(gmv, 2) if gmv is not None else None,
        "ebitda_usd":         round(ebt, 2) if ebt is not None else None,
        "ebitda_margin_pct":  em_pct,
    })
    stats[res] += 1

conn.commit()
print(f"  Consolidated View: inserted={stats['inserted']}  updated={stats['updated']}  "
      f"skipped={stats['skipped']}")
if cv_date_cols:
    periods_list = sorted(cv_date_cols.values())
    print(f"  Date range: {periods_list[0]} -> {periods_list[-1]}")


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 68)
print("AFTER — DETAILED SUMMARY")
print("=" * 68)
after = snapshot(COMPANY_IDS)

for cid in COMPANY_IDS:
    b = before[cid]
    a = after[cid]
    name = COMPANY_NAMES[cid]

    print(f"\n{'─'*50}")
    print(f"  {name} (company_id={cid})")
    print(f"{'─'*50}")
    print(f"  Rows:          {b['total']:3d} -> {a['total']:3d}")

    # Latest revenue
    latest = conn.execute(
        """
        SELECT period_end_date,
               revenue_usd,
               COALESCE(
                   ebitda_margin_pct,
                   CASE WHEN revenue_usd > 0 AND ebitda_usd IS NOT NULL
                        THEN ROUND(ebitda_usd * 100.0 / revenue_usd, 2)
                   END
               ) AS ebitda_pct,
               gross_margin_pct,
               loan_book_gross_usd,
               par_30_pct,
               net_yield_pct,
               arr_usd
        FROM kpi_snapshots
        WHERE company_id = ? AND revenue_usd IS NOT NULL
        ORDER BY period_end_date DESC
        LIMIT 1
        """,
        (cid,),
    ).fetchone()

    if latest:
        def _fmt(v):
            if v is None: return "—"
            if isinstance(v, float) and v > 1000: return f"${v:>12,.0f}"
            return f"{v:.2f}"

        print(f"  Latest period: {latest['period_end_date']}")
        print(f"  Revenue USD:   {_fmt(latest['revenue_usd'])}")
        print(f"  EBITDA margin: {_fmt(latest['ebitda_pct'])}%")
        print(f"  Gross margin:  {_fmt(latest['gross_margin_pct'])}%")
        print(f"  Loan book:     {_fmt(latest['loan_book_gross_usd'])}")
        print(f"  PAR 30:        {_fmt(latest['par_30_pct'])}%")
        print(f"  Net yield:     {_fmt(latest['net_yield_pct'])}%")
        print(f"  ARR:           {_fmt(latest['arr_usd'])}")
    else:
        print("  No revenue data found.")

    t = a["total"]
    if t > 0:
        print(f"\n  Metric coverage ({t} total rows):")
        for label, key in [
            ("revenue",      "rev"),
            ("gross_margin", "gm"),
            ("ebitda_usd",   "ebitda"),
            ("loan_book",    "loan"),
            ("par_30",       "par30"),
            ("net_yield",    "yield_"),
            ("arr_usd",      "arr"),
            ("customers",    "cust"),
        ]:
            filled = a[key]
            status = "OK" if filled == t else ("~" if filled > 0 else "XX")
            print(f"    {status} {label:15s} {filled:3d}/{t}")

    if b["total"] != a["total"]:
        print(f"\n  ▶ Added {a['total'] - b['total']} new rows")

conn.close()
print("\n" + "=" * 68)
print("Done.")
