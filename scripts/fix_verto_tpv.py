"""
Populate tpv_usd for Verto (company_id=3) from the board KPI Excel file.
Row 22 ("Total Processed Volume - Overall") contains monthly USD TPV.
Run from project root: python scripts/fix_verto_tpv.py
"""

import calendar
import sqlite3
from datetime import date

import openpyxl

DB_PATH   = "benchmarking.db"
XLSX_PATH = "data/Verto Monthly KPIs - 2026-01_Board.xlsx"

conn = sqlite3.connect(DB_PATH)

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb["KPIs"]

# Build date → column map from row 2
date_cols: dict[int, str] = {}
for c in range(3, ws.max_column + 1):
    raw = ws.cell(2, c).value
    if hasattr(raw, "year"):
        last_day = calendar.monthrange(raw.year, raw.month)[1]
        date_cols[c] = date(raw.year, raw.month, last_day).isoformat()

stats = {"updated": 0, "no_row": 0, "no_data": 0}
for col, period in sorted(date_cols.items(), key=lambda x: x[1]):
    tpv = ws.cell(22, col).value  # Total Processed Volume - Overall
    if tpv is None:
        stats["no_data"] += 1
        continue
    result = conn.execute(
        "UPDATE kpi_snapshots SET tpv_usd=?, updated_at=datetime('now') "
        "WHERE company_id=3 AND period_end_date=?",
        (float(tpv), period),
    )
    if result.rowcount > 0:
        stats["updated"] += 1
    else:
        stats["no_row"] += 1

conn.commit()

# Verify LTM TPV
ltm_tpv = conn.execute("""
    SELECT SUM(tpv_usd) FROM kpi_snapshots
    WHERE company_id=3 AND period_end_date BETWEEN '2025-02-01' AND '2026-01-31'
    AND tpv_usd IS NOT NULL
""").fetchone()[0]

conn.close()

print(f"Verto TPV: updated={stats['updated']}  no_row={stats['no_row']}  no_data={stats['no_data']}")
if ltm_tpv:
    print(f"LTM TPV (Feb 2025 – Jan 2026): ${ltm_tpv/1e9:.2f}B")
